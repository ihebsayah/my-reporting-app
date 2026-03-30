"""Document text extraction utilities for common source file types."""

import csv
import json
import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Sequence

logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class ExtractedDocument:
    """Represents normalized text extracted from a source document."""

    document_id: str
    text: str
    source_path: str
    file_type: str
    metadata: Dict[str, Any] = field(default_factory=dict)


class DocumentExtractionError(RuntimeError):
    """Raised when a source document cannot be normalized."""


class DocumentExtractor:
    """Extract normalized text from supported document types."""

    SUPPORTED_EXTENSIONS: Sequence[str] = (".txt", ".json", ".csv", ".pdf", ".xlsx")

    def extract_directory(self, input_dir: Path) -> List[ExtractedDocument]:
        """Extract all supported documents from a directory.

        Args:
            input_dir: Directory containing raw source documents.

        Returns:
            Extracted documents sorted by path name.

        Raises:
            FileNotFoundError: If the directory does not exist.
        """
        if not input_dir.exists():
            logger.error("Input directory does not exist: %s", input_dir)
            raise FileNotFoundError(f"Input directory does not exist: {input_dir}")

        documents: List[ExtractedDocument] = []
        for path in sorted(input_dir.iterdir()):
            if path.is_dir() or path.suffix.lower() not in self.SUPPORTED_EXTENSIONS:
                continue
            documents.append(self.extract_file(path))

        logger.info("Extracted %d documents from %s.", len(documents), input_dir)
        return documents

    def extract_file(self, path: Path) -> ExtractedDocument:
        """Extract a single supported file.

        Args:
            path: Source file path.

        Returns:
            Normalized extracted document.

        Raises:
            DocumentExtractionError: If the extension is unsupported or parsing fails.
        """
        suffix = path.suffix.lower()
        logger.info("Extracting text from %s.", path)

        try:
            if suffix == ".txt":
                return self._extract_txt(path)
            if suffix == ".json":
                return self._extract_json(path)
            if suffix == ".csv":
                return self._extract_csv(path)
            if suffix == ".pdf":
                return self._extract_pdf(path)
            if suffix == ".xlsx":
                return self._extract_xlsx(path)
        except Exception as exc:
            logger.error("Failed to extract document %s: %s", path, exc)
            raise DocumentExtractionError(
                f"Failed to extract document {path}: {exc}"
            ) from exc

        logger.error("Unsupported file type for extraction: %s", path)
        raise DocumentExtractionError(f"Unsupported file type: {path.suffix}")

    def _extract_txt(self, path: Path) -> ExtractedDocument:
        """Extract text from a plain text file."""
        text = path.read_text(encoding="utf-8")
        return ExtractedDocument(
            document_id=path.stem,
            text=text,
            source_path=str(path),
            file_type="txt",
            metadata={"source_name": path.name},
        )

    def _extract_json(self, path: Path) -> ExtractedDocument:
        """Extract text from a JSON document payload."""
        payload = json.loads(path.read_text(encoding="utf-8"))
        if not isinstance(payload, dict) or "text" not in payload:
            raise ValueError("JSON document must be an object containing 'text'.")

        document_id = payload.get("document_id", path.stem)
        metadata = payload.get("metadata", {})
        if not isinstance(document_id, str):
            raise ValueError("JSON document_id must be a string.")
        if not isinstance(metadata, dict):
            raise ValueError("JSON metadata must be an object.")

        metadata = {
            **metadata,
            "source_name": path.name,
        }
        return ExtractedDocument(
            document_id=document_id,
            text=str(payload["text"]),
            source_path=str(path),
            file_type="json",
            metadata=metadata,
        )

    def _extract_csv(self, path: Path) -> ExtractedDocument:
        """Extract text from a CSV file by flattening rows into labeled lines."""
        rows: List[str] = []
        with path.open("r", encoding="utf-8", newline="") as handle:
            reader = csv.reader(handle)
            for row_number, row in enumerate(reader, start=1):
                row_text = " | ".join(cell.strip() for cell in row)
                rows.append(f"row_{row_number}: {row_text}")

        return ExtractedDocument(
            document_id=path.stem,
            text="\n".join(rows),
            source_path=str(path),
            file_type="csv",
            metadata={"source_name": path.name, "row_count": len(rows)},
        )

    def _extract_pdf(self, path: Path) -> ExtractedDocument:
        """Extract text from a PDF file using an optional dependency."""
        try:
            from pypdf import PdfReader  # type: ignore
        except ImportError as exc:
            raise ImportError("pypdf is required for PDF extraction.") from exc

        reader = PdfReader(str(path))
        pages = [page.extract_text() or "" for page in reader.pages]
        text = "\n".join(page.strip() for page in pages if page.strip())
        return ExtractedDocument(
            document_id=path.stem,
            text=text,
            source_path=str(path),
            file_type="pdf",
            metadata={"source_name": path.name, "page_count": len(reader.pages)},
        )

    def _extract_xlsx(self, path: Path) -> ExtractedDocument:
        """Extract text from an Excel workbook using an optional dependency."""
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError as exc:
            raise ImportError("openpyxl is required for Excel extraction.") from exc

        workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
        sheet_text: List[str] = []
        for worksheet in workbook.worksheets:
            sheet_text.append(f"[sheet:{worksheet.title}]")
            for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                values = ["" if value is None else str(value).strip() for value in row]
                if any(values):
                    sheet_text.append(f"row_{row_number}: {' | '.join(values)}")

        return ExtractedDocument(
            document_id=path.stem,
            text="\n".join(sheet_text),
            source_path=str(path),
            file_type="xlsx",
            metadata={
                "source_name": path.name,
                "sheet_count": len(workbook.worksheets),
            },
        )


def build_annotation_documents(
    extracted_documents: List[ExtractedDocument],
) -> List[Dict[str, Any]]:
    """Convert extracted documents into annotation-ready dictionaries.

    Args:
        extracted_documents: Normalized extracted documents.

    Returns:
        Annotation-ready dictionaries carrying normalized text and source metadata.
    """
    payloads: List[Dict[str, Any]] = []
    for document in extracted_documents:
        payloads.append(
            {
                "document_id": document.document_id,
                "text": document.text,
                "metadata": {
                    **document.metadata,
                    "source_path": document.source_path,
                    "file_type": document.file_type,
                },
            }
        )
    return payloads
